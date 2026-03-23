from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import load_workbook
except ImportError:  # optional dependency for xlsx support
    load_workbook = None


@dataclass(frozen=True)
class SupplierProjectRecord:
    project_name: str
    supplier_name: str


HEADER_PROJECT = "项目名称"
HEADER_SUPPLIER = "供应商名称"


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_rows(rows: List[List[object]]) -> List[SupplierProjectRecord]:
    if not rows:
        return []

    headers = [_clean_cell(cell) for cell in rows[0]]
    try:
        project_idx = headers.index(HEADER_PROJECT)
        supplier_idx = headers.index(HEADER_SUPPLIER)
    except ValueError as exc:
        raise ValueError(
            f"模板表头缺失，必须包含“{HEADER_PROJECT}”和“{HEADER_SUPPLIER}”"
        ) from exc

    records: List[SupplierProjectRecord] = []
    for row in rows[1:]:
        project_name = _clean_cell(row[project_idx] if project_idx < len(row) else "")
        supplier_name = _clean_cell(row[supplier_idx] if supplier_idx < len(row) else "")
        if not project_name and not supplier_name:
            continue
        if not supplier_name:
            continue
        records.append(SupplierProjectRecord(project_name=project_name, supplier_name=supplier_name))

    return records


def _read_xlsx(template_path: Path) -> List[List[object]]:
    if load_workbook is None:
        raise RuntimeError("未安装 openpyxl，无法读取 xlsx 模板，请先执行 pip install openpyxl")

    workbook = load_workbook(filename=str(template_path), read_only=True, data_only=True)
    sheet = workbook.active
    rows: List[List[object]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    workbook.close()
    return rows


def _read_csv(template_path: Path) -> List[List[object]]:
    with template_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        return [list(row) for row in reader]


def load_supplier_project_records(template_path: str) -> List[SupplierProjectRecord]:
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"模板文件不存在: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = _read_xlsx(path)
    elif suffix == ".csv":
        rows = _read_csv(path)
    else:
        raise ValueError("仅支持 .xlsx 或 .csv 模板文件")

    records = _parse_rows(rows)
    if not records:
        raise ValueError("模板中未读取到有效供应商数据")
    return records


def group_suppliers_by_project(records: List[SupplierProjectRecord]) -> Dict[str, List[str]]:
    grouped: Dict[str, List[str]] = {}
    for record in records:
        project_name = record.project_name or "未命名项目"
        supplier_list = grouped.setdefault(project_name, [])
        if record.supplier_name not in supplier_list:
            supplier_list.append(record.supplier_name)
    return grouped
